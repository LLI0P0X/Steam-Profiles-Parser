import openpyxl
import os
import sys

if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)
elif __file__:
    application_path = os.path.dirname(__file__)

os.chdir(application_path)
path = os.path.join(application_path)

if not os.path.exists(os.path.join(path, 'out')):
    os.makedirs(os.path.join(path, 'out'))

pathTemp = os.path.join(path, 'temp')
pathOut = os.path.join(path, 'out')


def flt(st):
    return str.isspace(st) or str.isalpha(st) or str.isalnum(st)


def safeName(name):
    return ''.join(filter(flt, name)).replace(' ', '_')


gbl = globals()
row = 0


def toExcel(data, name):
    name = safeName(name)
    gbl = globals()
    row = gbl['row']
    if row:
        wb = openpyxl.load_workbook(os.path.join(pathOut, f'{name}.xlsx'))
        ws = wb.get_sheet_by_name('data')
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('data')
        wb.remove(wb.get_sheet_by_name('Sheet'))
    for col in range(len(data)):
        d = data[col]
        if len(d) == 0:
            continue
        if d[0] == '=':
            d = "'" + d
        ws.cell(row + 1, col + 1, d)
    gbl['row'] += 1
    wb.save(os.path.join(pathOut, f'{name}.xlsx'))
    return f'{os.path.join(pathOut, name)}.xlsx'


if __name__ == '__main__':
    pass
